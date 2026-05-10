"""
建立 / 更新股癌節目的股票追蹤資料庫。

格式：
- 工作表「股票清單」：去重彙總，記錄首次提到、最近提到、累計次數、觀點摘要
- 工作表「逐集記錄」：每一集提到的股票完整明細

使用方式：
    python build_gooaye_stocks.py            # 重建（從種子資料）
    python build_gooaye_stocks.py --append "EP659" "2026-05-08" "AAPL,蘋果,美股,觀點"
"""
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DIR = Path(__file__).resolve().parent.parent / "report"
XLSX_PATH = REPORT_DIR / "股癌_股票追蹤.xlsx"

# 種子資料：EP657, EP658 提到的股票
# 欄位：(代號, 中文/英文名, 市場, 集數, 日期, 觀點摘要)
SEED_DATA = [
    # ========== EP657 ==========
    ("2330", "台積電", "台股", "EP657", "2026-04-25", "投信進來買，外資死命賣（一天 1.8~2 萬張）；2330 元上限不在這裡，但籌碼面承壓"),
    ("2454", "聯發科", "台股", "EP657", "2026-04-25", "Goldman Sachs 喊目標價 2454 元；MTK 強勢"),
    ("2337", "旺宏", "台股", "EP657", "2026-04-25", "老闆法說會轉為超級利多；MLC NAND / Nord 漲價明顯"),
    ("6531", "愛普", "台股", "EP657", "2026-04-25", "做 Intel 載板裡的 IPD（埋入載板的被動式電容）"),
    ("INTC", "Intel", "美股", "EP657", "2026-04-25", "後段封裝（assembly）良率拉到 90%，可量產；前段仍困難"),
    ("AMZN", "Amazon", "美股", "EP657", "2026-04-25", "AWS Graviton5 與 Meta 簽 Multi-billion 協議；Graviton 供不應求"),
    ("META", "Meta", "美股", "EP657", "2026-04-25", "簽 AWS Graviton5 大型部署；CPU 部署激進，需 Tens of Millions 個 Cores"),
    ("AMD", "AMD", "美股", "EP657", "2026-04-25", "CPU 火熱供應鏈受惠"),
    ("ARM", "Arm Holdings", "美股", "EP657", "2026-04-25", "CPU 供應鏈相關"),
    ("QCOM", "高通", "美股", "EP657", "2026-04-25", "與 MediaTek 同為 OpenAI 手機晶片供應商，每顆 $120 美元"),
    ("MTK_OPENAI", "OpenAI x 力訊精密", "未上市", "EP657", "2026-04-25", "OpenAI 跟力訊精密做原生 AI 消費設備（Smartphone）；下下覺得難挑戰 iPhone"),
    ("0050", "元大台灣50", "ETF", "EP657", "2026-04-25", "推薦長期持有的核心部位（搭配 006208 / VOO / VTI）"),
    ("006208", "富邦台50", "ETF", "EP657", "2026-04-25", "推薦的核心 ETF 之一"),
    ("00713", "元大高息低波", "ETF", "EP657", "2026-04-25", "聽眾提問配 0050 + 00713 質押維持率管理"),
    ("VOO", "Vanguard S&P 500", "美股 ETF", "EP657", "2026-04-25", "推薦的長期持有 ETF"),
    ("VTI", "Vanguard Total Stock Market", "美股 ETF", "EP657", "2026-04-25", "推薦的長期持有 ETF"),
    ("ARKK", "ARK Innovation ETF", "美股 ETF", "EP657", "2026-04-25", "歷史教訓：2021 暴利後一蹶不振，木頭姐尚未站起來"),
    # ========== EP658 ==========
    ("2330", "台積電", "台股", "EP658", "2026-04-28", "AI 世界唯一沒有對手；製程 16A→14A→12A；CoWoS / SoIC 雙封裝"),
    ("2454", "聯發科", "台股", "EP658", "2026-04-28", "TSMC 先進封裝大將加入 MTK；TPU 5922 後段回歸台積機率升高；發哥法說 EPS 估近 400"),
    ("INTC", "Intel", "美股", "EP658", "2026-04-28", "好消息環繞但歷史常粗暴收場；EMIB 良率帶來機會但不宜追加"),
    ("TEAM", "Atlassian", "美股", "EP658", "2026-04-28", "財報意外亮眼，Seat expansion 持續；打臉 SaaS Apocalypse"),
    ("CRM", "Salesforce", "美股", "EP658", "2026-04-28", "中性偏好；Founder-led（Benioff），AgentForce + Headless 架構"),
    ("ADBE", "Adobe", "美股", "EP658", "2026-04-28", "觀望；初階應用被 AI 搶走，估值打壓合理"),
    ("PLTR", "Palantir", "美股", "EP658", "2026-04-28", "最看好之一；FDE 客製化部署 + 垂直整合難被 AI 取代"),
    ("CRWD", "CrowdStrike", "美股", "EP658", "2026-04-28", "最抗 AI 替代的軟體股之一"),
    ("NET", "Cloudflare", "美股", "EP658", "2026-04-28", "最抗 AI 替代的軟體股之一"),
    ("CHKP", "Check Point", "美股", "EP658", "2026-04-28", "原被市場判死刑（與 Atlassian 同列），意外活過來"),
    ("GOOGL", "Alphabet (Google)", "美股", "EP658", "2026-04-28", "Google TPU 自研晶片 vs NVIDIA GPU；不是零和遊戲；TPU 外銷視為 Cloud 策略"),
    ("AMZN", "Amazon", "美股", "EP658", "2026-04-28", "自研 Trainium 與採購 GPU 並行"),
    ("NVDA", "NVIDIA", "美股", "EP658", "2026-04-28", "黃仁勳訪問：AI 不會取代『人』，Coding 前所未有重要；推進人型機器人"),
    ("AMD", "AMD", "美股", "EP658", "2026-04-28", "AI 晶片整體供不應求，與 NVIDIA 並非零和"),
    ("MU", "美光 Micron", "美股", "EP658", "2026-04-28", "美光下去時一堆人說破底做頭，結果直接屌噴"),
    ("SNDK", "SanDisk", "美股", "EP658", "2026-04-28", "與美光類似，下殺後屌噴"),
    ("AAPL", "Apple", "美股", "EP658", "2026-04-28", "電話會議再次確認後續記憶體成本將持續上漲"),
    ("MSFT", "Microsoft", "美股", "EP658", "2026-04-28", "Google 強之後不會再有人嘴 Search；巨頭 AI 並未殺死軟體"),
    ("TSLA", "Tesla", "美股", "EP658", "2026-04-28", "與 NVIDIA 持續推進人型機器人"),
]


def build():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # 主表：股票清單（去重彙總）
    ws_main = wb.active
    ws_main.title = "股票清單"
    headers_main = ["代號", "名稱", "市場", "首次提到集數", "首次提到日期", "最近提到集數", "最近提到日期", "累計提到次數", "觀點摘要（最新）"]
    ws_main.append(headers_main)

    # 逐集記錄
    ws_log = wb.create_sheet("逐集記錄")
    headers_log = ["集數", "日期", "代號", "名稱", "市場", "觀點摘要"]
    ws_log.append(headers_log)

    aggregated = {}  # key: (代號, 名稱)
    for ticker, name, market, ep, date, note in SEED_DATA:
        ws_log.append([ep, date, ticker, name, market, note])
        key = (ticker, name)
        if key not in aggregated:
            aggregated[key] = {
                "market": market,
                "first_ep": ep,
                "first_date": date,
                "last_ep": ep,
                "last_date": date,
                "count": 1,
                "note": note,
            }
        else:
            agg = aggregated[key]
            agg["count"] += 1
            if date > agg["last_date"]:
                agg["last_ep"] = ep
                agg["last_date"] = date
                agg["note"] = note

    # 排序：首次提到越早越前
    sorted_items = sorted(aggregated.items(), key=lambda x: (x[1]["first_date"], x[0][0]))
    for (ticker, name), agg in sorted_items:
        ws_main.append([
            ticker, name, agg["market"],
            agg["first_ep"], agg["first_date"],
            agg["last_ep"], agg["last_date"],
            agg["count"], agg["note"],
        ])

    # 樣式
    header_font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    body_font = Font(name="Microsoft JhengHei", size=10)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for ws in (ws_main, ws_log):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = align_left
                cell.border = border
        ws.freeze_panes = "A2"

    # 欄寬
    widths_main = [12, 22, 12, 14, 14, 14, 14, 14, 60]
    for i, w in enumerate(widths_main, start=1):
        ws_main.column_dimensions[get_column_letter(i)].width = w
    widths_log = [10, 12, 12, 22, 12, 70]
    for i, w in enumerate(widths_log, start=1):
        ws_log.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_PATH)
    print(f"[OK] Saved: {XLSX_PATH}")
    print(f"  - 股票清單: {len(aggregated)} 檔股票")
    print(f"  - 逐集記錄: {len(SEED_DATA)} 筆")


def append_episode(ep, date, items):
    """
    新增一集的股票提及記錄。
    items: list of (代號, 名稱, 市場, 觀點摘要)
    """
    if not XLSX_PATH.exists():
        print("[ERROR] 資料庫尚未建立，請先執行 build()")
        return

    wb = load_workbook(XLSX_PATH)
    ws_main = wb["股票清單"]
    ws_log = wb["逐集記錄"]

    # 建立既存彙總索引
    existing = {}
    for row in ws_main.iter_rows(min_row=2, values_only=False):
        ticker = row[0].value
        name = row[1].value
        if ticker is None:
            continue
        existing[(ticker, name)] = row

    for ticker, name, market, note in items:
        ws_log.append([ep, date, ticker, name, market, note])
        key = (ticker, name)
        if key in existing:
            row = existing[key]
            # 更新最近提到 + 累計次數 + 觀點
            row[5].value = ep        # 最近提到集數
            row[6].value = date      # 最近提到日期
            row[7].value = (row[7].value or 0) + 1
            row[8].value = note
        else:
            ws_main.append([ticker, name, market, ep, date, ep, date, 1, note])

    wb.save(XLSX_PATH)
    print(f"[OK] 已附加 {ep} ({date})：{len(items)} 筆")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--append":
        # python build_gooaye_stocks.py --append EP659 2026-05-08 "AAPL|蘋果|美股|觀點摘要;NVDA|輝達|美股|觀點摘要"
        ep = sys.argv[2]
        date = sys.argv[3]
        raw = sys.argv[4]
        items = []
        for chunk in raw.split(";"):
            parts = chunk.split("|")
            if len(parts) == 4:
                items.append(tuple(parts))
        append_episode(ep, date, items)
    else:
        build()
